import os.path

import openpyxl
from openpyxl import load_workbook

from tender.utils import date
from tender.common.consts import const


class ExcelPipeline:
    def __init__(self):
        self.wb = None
        self.sheet = None
        self.keywordsDict = dict()
        self.webnamesDict = dict()
        self.getWebnames()

    def open_spider(self, spider):
        self.wb = openpyxl.Workbook()
        sheet_name = self.webnamesDict.get(spider.name)
        self.sheet = self.wb.create_sheet(sheet_name,0)
        self.sheet.append(['标题', '创建时间', 'URL'])

    def close_spider(self, spider):
        print("Excel数据保存结束---------------------")
        self.wb.save('excel/tender-{}-{}.xlsx'.format(date.get_curdate(),spider.name))

    def process_item(self, item, spider):
        pubtime = item['pubtime']
        title = item['title']
        self.getKeywords()
        keywordStr = self.keywordsDict.get(spider.name)
        if self.checkTitle(title, keywordStr) and date.get_curdate() == pubtime:
            data = [item['title'], item['pubtime'], item['link']]
            self.sheet.append(data)
        return item
    def checkTitle(self, title, keywordStr):
        title = str(title)
        keywords = keywordStr.split(",")
        t = []
        for k in keywords:
            if k in title:
                t.append(True)
        if len(t) > 0:
            return True
        else:
            return False

    def getKeywords(self):
        # 从配置文件中读取关键字
        curpath = os.getcwd()
        fk = open(curpath + '\\' + const.KEYWORD_CONF, 'r', encoding='utf-8')
        while True:
            line = fk.readline()
            if line:
                config = line.split('===')
                self.keywordsDict[config[0]] = config[1]
            else:
                break
        fk.close()

    def getWebnames(self):
        # 从配置文件中读取关键字
        curpath = os.getcwd()
        fw = open(curpath + "\\" + const.WEBNAME_CONF, 'r', encoding='utf-8')
        while True:
            line = fw.readline()
            if line:
                config = line.split('===')
                self.webnamesDict.setdefault(config[0], config[1])
            else:
                break
        fw.close()